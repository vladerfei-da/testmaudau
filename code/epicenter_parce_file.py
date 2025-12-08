import os
from bs4 import BeautifulSoup
import csv
import re

class EpicentrkParser:
    """
    Парсер HTML-сторінок каталогу Epicentr K для локальних файлів.
    """
    def __init__(self):
        self.base_url = "https://epicentrk.ua"
        self.products = []
    
    def read_local_page(self, filepath):
        """Отримати HTML з локального файлу"""
        try:
            print(f"Завантаження: {filepath}")
            with open(filepath, 'r', encoding='utf-8') as f:
                html = f.read()
            return html
        except FileNotFoundError:
            print(f"❌ Помилка: Файл не знайдено за шляхом: {filepath}")
            return None
        except Exception as e:
            print(f"❌ Помилка читання файлу {filepath}: {e}")
            return None
    
    def parse_page_html(self, html):
        """Парсинг сторінки через BeautifulSoup для Epicentr K."""
        soup = BeautifulSoup(html, 'html.parser')
        products_found = []
        
        product_items = soup.find_all('li', {'data-test-small-card': re.compile(r'\d+')})
        
        print(f"   (DEBUG: Знайдено {len(product_items)} блоків товарів)")

        for container in product_items:
            try:
                # 1. Назва та URL
                link_elem = container.find('a', {'itemprop': 'url'})
                if not link_elem:
                    continue 

                product_url = link_elem.get('href', '')
                url = f"{self.base_url}{product_url}" if product_url.startswith('/') else product_url
                
                title = link_elem.get('title', '') or link_elem.get_text(strip=True)
                if not title:
                    continue
                    
                # 2. Зображення
                image = ""
                img = container.find('img', {'itemprop': 'image'})
                if img:
                    image = img.get('src', '') or img.get('data-src', '')
                    
                # 3. Актуальна ціна (data-product-price-main)
                price = ""
                price_elem = container.find('div', {'data-product-price-main': True})
                if price_elem:
                    # Шукаємо перший тег <data> з value
                    data_tag = price_elem.find('data', {'value': True})
                    if data_tag:
                        price_val = data_tag.get('value', '0')
                        price = str(int(float(price_val)))
                
                # 4. Стара ціна (s data-product-price-old)
                old_price = ""
                old_price_elem = container.find('s', {'data-product-price-old': True})
                if old_price_elem:
                    # Шукаємо тег <data> з атрибутом content
                    data_tag = old_price_elem.find('data', {'content': True})
                    if data_tag:
                        old_price_val = data_tag.get('content', '0')
                        old_price = str(int(float(old_price_val)))
                
                # 5. Знижка в гривнях (small data-product-price-badge)
                discount_amount_uah = ""
                discount_badge_elem = container.find('small', {'data-product-price-badge': True})
                if discount_badge_elem:
                    # Шукаємо перший тег <data> з числовим value
                    data_tags = discount_badge_elem.find_all('data', {'value': True})
                    for data_tag in data_tags:
                        val = data_tag.get('value', '')
                        # Перевіряємо, чи це число (не "UAH")
                        if val and val.replace('.', '').replace(',', '').isdigit():
                            discount_amount_uah = str(int(float(val)))
                            break
                    
                # 6. Відсоток знижки (розрахунок)
                discount_percent = ""
                if price and old_price:
                    try:
                        price_int = int(price)
                        old_price_int = int(old_price)
                        if old_price_int > price_int:
                            percent = round((old_price_int - price_int) / old_price_int * 100)
                            discount_percent = str(percent)
                    except ValueError:
                        pass
                
                # 7. Рейтинг та відгуки (відсутні на картці)
                rating = "" 
                reviews = ""

                # 8. Бренд (з dl/dt/dd структури)
                brand = ""
                # Шукаємо <dt itemprop="name">Бренд</dt>
                brand_dt = container.find('dt', string='Бренд')
                if brand_dt:
                    # Наступний елемент повинен бути <dd itemprop="value">
                    brand_dd = brand_dt.find_next_sibling('dd')
                    if brand_dd:
                        brand = brand_dd.get_text(strip=True)
                
                # Якщо бренд не знайдено, беремо перше слово з назви
                if not brand and title:
                    brand = title.split()[0]
                    
                # 9. Наявність
                availability_text = container.find(lambda tag: tag.name == 'span' and 'Немає в наявності' in tag.get_text(strip=True))
                availability = "Немає в наявності" if availability_text else "В наявності" 
                
                # 10. Бейджі (стікери знижок)
                badges = ""
                sticker = container.find('div', {'data-sticker-title': True})
                if sticker:
                    badges = sticker.get('data-sticker-title', '')

                if title and price:
                    products_found.append({
                        'title': title,
                        'url': url,
                        'image': image,
                        'price': price,
                        'old_price': old_price,
                        'discount_amount_uah': discount_amount_uah,
                        'rating': rating,
                        'reviews_count': reviews,
                        'availability': availability,
                        'badges': badges,
                        'brand': brand,
                        'discount_percent': discount_percent
                    })
                    
            except Exception as e:
                print(f"   ⚠️ Помилка парсингу товару: {e}")
                continue
        
        return products_found

    def parse_file(self, filepath):
        """Парсинг одного локального файлу"""
        html = self.read_local_page(filepath)
        if not html:
            return False
        
        print(f"🔄 Парсинг через HTML...")
        page_products = self.parse_page_html(html)
        
        if not page_products:
            print("⛔ Не вдалося знайти товари.")
            return False

        print(f"✅ Знайдено {len(page_products)} товарів")
        
        discount_count = 0
        for p in page_products:
            price_info = f"{p['price']} ₴"
            if p.get('discount_amount_uah') or p.get('old_price'):
                discount_count += 1
                if p.get('discount_amount_uah'):
                    price_info = f"Ціна: {p['price']} ₴ | Економія: {p['discount_amount_uah']} ₴"
                elif p.get('old_price'):
                    price_info = f"{p['old_price']} ₴ → {p['price']} ₴"

            if p['discount_percent']:
                price_info += f" (-{p['discount_percent']}%)"
            
            print(f"   ✓ {p['title'][:55]}... | {price_info}")
        
        print(f"   (Акційних товарів: {discount_count})")
        
        self.products.extend(page_products)
        return True
    
    def parse_all_local_files(self, folder_path):
        """Парсинг всіх HTML файлів у папці"""
        
        # ✅ ПРАВИЛЬНО: Сканування папки
        if not os.path.exists(folder_path):
            print(f"❌ Папка не існує: {folder_path}")
            return
        
        # Отримуємо список HTML файлів
        html_files = [f for f in os.listdir(folder_path) if f.endswith('.html')]
        html_files.sort()  # Сортуємо для порядку
        
        total_files = len(html_files)
        
        print(f"\n{'='*80}")
        print(f"🚀 ПОЧАТОК ЛОКАЛЬНОГО ПАРСИНГУ EPICENTRK.UA")
        print(f"📂 Папка: {folder_path}")
        print(f"📄 Знайдено HTML файлів: {total_files}")
        print(f"{'='*80}\n")
        
        if total_files == 0:
            print("❌ HTML файли не знайдено!")
            return
        
        # Парсимо кожен файл
        for i, filename in enumerate(html_files, 1):
            filepath = os.path.join(folder_path, filename)
            print(f"\n[{i}/{total_files}] Обробка: {filename}")
            self.parse_file(filepath)
            
        print(f"\n✅ Зібрано: {len(self.products)} товарів (загалом)")
        print(f"\n{'='*80}")
        print(f"🎉 ПАРСИНГ ЗАВЕРШЕНО!")
        print(f"{'='*80}")

    def save_to_csv(self, filename='epik_krasa_fluid_oblychchia.csv'):
        """Збереження у CSV"""
        if not self.products:
            print("❌ Немає даних для збереження")
            return
        
        with open(filename, 'w', newline='', encoding='utf-8-sig') as f:
            fields = ['title', 'url', 'image', 'price', 'old_price', 'discount_amount_uah', 'rating',
                      'reviews_count', 'availability', 'badges', 'brand', 'discount_percent']
            writer = csv.DictWriter(f, fieldnames=fields)
            writer.writeheader()
            writer.writerows(self.products)
        
        print(f"\n✅ CSV збережено: {filename}")
        print(f"📊 Товарів: {len(self.products)}")
    
    def save_to_excel(self, filename='epik_krasa_fluid_oblychchia.xlsx'):
        """Збереження у Excel"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Товари"
            
            headers = ['№', 'Назва', 'URL', 'Зображення', 'Ціна', 'Стара ціна', 'Знижка (₴)',
                       'Рейтинг', 'Відгуків', 'Наявність', 'Бейджі', 'Бренд', 'Знижка %']
            
            fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            font = Font(bold=True, color="FFFFFF", size=11)
            
            for col, h in enumerate(headers, 1):
                cell = ws.cell(1, col, h)
                cell.fill = fill
                cell.font = font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            for idx, p in enumerate(self.products, 2):
                ws.cell(idx, 1, idx-1)
                ws.cell(idx, 2, p['title'])
                ws.cell(idx, 3, p['url'])
                ws.cell(idx, 4, p['image'])
                ws.cell(idx, 5, p['price'])
                ws.cell(idx, 6, p.get('old_price', ''))
                ws.cell(idx, 7, p.get('discount_amount_uah', ''))
                ws.cell(idx, 8, p['rating'])
                ws.cell(idx, 9, p['reviews_count'])
                ws.cell(idx, 10, p['availability'])
                ws.cell(idx, 11, p['badges'])
                ws.cell(idx, 12, p['brand'])
                ws.cell(idx, 13, p['discount_percent'])
            
            for col in ws.columns:
                max_len = 0
                col_letter = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_len:
                            max_len = len(cell.value)
                    except:
                        pass
                ws.column_dimensions[col_letter].width = min(max_len + 2, 60)
            
            wb.save(filename)
            print(f"✅ Excel збережено: {filename}")
            
        except ImportError:
            print("⚠️ Встановіть openpyxl: pip install openpyxl")

    def print_sample(self, count=5):
        """Показати приклад товарів"""
        if not self.products:
            print("❌ Немає товарів для відображення")
            return
        
        print(f"\n{'='*80}")
        print(f"📋 ПРИКЛАД ТОВАРІВ ({min(count, len(self.products))} шт.):")
        print(f"{'='*80}\n")
        
        for i, p in enumerate(self.products[:count], 1):
            print(f"{i}. {p['title']}")
            print(f"   🏷️  Бренд: {p['brand']}")
            if p.get('old_price'):
                print(f"   💰 Стара: {p['old_price']} ₴")
            print(f"   💵 Ціна: {p['price']} ₴")
            if p.get('discount_amount_uah'):
                print(f"   💾 Економія: {p['discount_amount_uah']} ₴")
            if p['discount_percent']:
                print(f"   🔥 Знижка: -{p['discount_percent']}%")
            print(f"   📦 {p['availability']}")
            print(f"   🔗 {p['url']}\n")


if __name__ == "__main__":
    # ✅ ПРАВИЛЬНІ ШЛЯХИ
    HTML_FOLDER_PATH = r'D:\testmaudau\epik_krasa_fluid_oblychchia'
    OUTPUT_FOLDER = r'D:\testmaudau'
    
    parser = EpicentrkParser()
    
    # Парсинг всіх HTML-файлів з папки
    parser.parse_all_local_files(folder_path=HTML_FOLDER_PATH)
    
    # Показати приклади
    parser.print_sample(5)
    
    # Зберегти результати
    parser.save_to_csv(os.path.join(OUTPUT_FOLDER, 'epik_krasa_fluid_oblychchia.csv'))
    parser.save_to_excel(os.path.join(OUTPUT_FOLDER, 'epik_krasa_fluid_oblychchia.xlsx'))
    
    print("\n✨ Готово!")