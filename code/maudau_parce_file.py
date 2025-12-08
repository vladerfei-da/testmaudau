import os
from bs4 import BeautifulSoup
import csv
import json
import re
# requests та time не потрібні для локального парсингу, але залишаємо їх умовним імпортом
# import requests 
# import time 

class MaudauParser:
    """
    Парсер HTML-сторінок каталогу Maudau, адаптований для 
    читання файлів з локальної папки. Використовує гібридний 
    метод парсингу: спочатку JSON, потім посилений HTML (data-testid).
    """
    def __init__(self):
        self.base_url = "https://maudau.com.ua"
        self.products = []
        # self.session = requests.Session() # Не використовується для локального парсингу
    
    ## 🛠️ Методи для роботи з файлами
    
    def read_local_page(self, filepath):
        """Отримати HTML з локального файлу"""
        try:
            print(f"Завантаження: {filepath}")
            # Використовуємо 'r' та 'utf-8' для читання HTML-файлу
            with open(filepath, 'r', encoding='utf-8') as f:
                html = f.read()
            return html
        except FileNotFoundError:
            print(f"❌ Помилка: Файл не знайдено за шляхом: {filepath}")
            return None
        except Exception as e:
            print(f"❌ Помилка читання файлу {filepath}: {e}")
            return None
    
    ## 🧬 Методи вилучення даних
    
    def extract_json_data(self, html):
        """Витягнути JSON дані з HTML (пошук основного блоку даних)"""
        json_products = []
        
        # Спроба 1: Пошук масиву продуктів 
        pattern_array = r'"products":(\[.*?\])'
        match_products = re.search(pattern_array, html, re.DOTALL)

        if match_products:
            try:
                products_array_str = match_products.group(1)
                products_array = json.loads(products_array_str)
                json_products.extend(products_array) 
            except json.JSONDecodeError:
                pass
                
        # Спроба 2: Пошук окремих об'єктів "product"
        if not json_products:
            pattern_single = r'"product":\s*(\{(?:[^{}]|\{(?:[^{}]|\{[^{}]*\})*\})*\})'
            matches = re.finditer(pattern_single, html)
            
            for match in matches:
                try:
                    product_data = json.loads(match.group(1))
                    json_products.append(product_data)
                except json.JSONDecodeError:
                    continue
        
        # Фільтруємо, залишаючи лише ті, що мають ціну
        json_products = [p for p in json_products if p.get('offer', {}).get('price')]
        
        return json_products
    
    def parse_product_from_json(self, product_data):
        """Парсинг товару з JSON даних"""
        try:
            title = product_data.get('title', '')
            slug = product_data.get('slug', '')
            url = f"{self.base_url}/product/{slug}"
            
            main_photo = product_data.get('main_photo_sized_urls', {})
            image = main_photo.get('lg', '') or main_photo.get('md', '') or main_photo.get('xl', '')
            
            # Ціни та знижки (в копійках)
            offer = product_data.get('offer', {})
            price_cents = offer.get('price', 0)
            old_price_cents_raw = offer.get('old_price', 0)
            discount_amount_cents = offer.get('discount_amount', 0)
            
            # Конвертуємо в гривні
            price = str(price_cents // 100) if price_cents else ""
            
            discount_amount_uah = ""
            if discount_amount_cents and discount_amount_cents > 0:
                discount_amount_uah = str(discount_amount_cents // 100)
            
            old_price = ""
            if old_price_cents_raw and old_price_cents_raw > price_cents:
                old_price = str(old_price_cents_raw // 100)
            elif discount_amount_uah and price and discount_amount_cents > 0:
                try:
                    # Розрахунок старої ціни, якщо є тільки знижка та актуальна ціна
                    calculated_old_price = (price_cents + discount_amount_cents) // 100
                    old_price = str(calculated_old_price)
                except:
                    pass 
            
            discount_percent = str(offer.get('discount_percentage', '')) if offer.get('discount_percentage') else ""
            rating = str(product_data.get('rating', '')) if product_data.get('rating') else ""
            reviews_count = str(product_data.get('reviews_count', '')) if product_data.get('reviews_count') else ""
            
            available = offer.get('available', False)
            stock = offer.get('stock', 0)
            availability = "Немає в наявності"
            if available:
                if stock > 0:
                    availability = "В наявності"
                elif stock == 0:
                    availability = "Очікується"
            
            badges_list = product_data.get('badges', [])
            if isinstance(badges_list, list):
                badges = ', '.join(str(b) for b in badges_list if b) if badges_list else ""
            else:
                badges = str(badges_list) if badges_list else ""
            
            brand_data = product_data.get('brand', {})
            if isinstance(brand_data, dict):
                brand = brand_data.get('name', '') or brand_data.get('slug', '')
            else:
                brand = str(brand_data) if brand_data else ""
            
            if discount_amount_uah or old_price:
                 print(f"   (DEBUG: JSON: Знижка: {discount_amount_uah} ₴ | Стара ціна: {old_price} ₴)")
            
            return {
                'title': title,
                'url': url,
                'image': image,
                'price': price,
                'old_price': old_price, 
                'discount_amount_uah': discount_amount_uah,
                'rating': rating,
                'reviews_count': reviews_count,
                'availability': availability,
                'badges': badges,
                'brand': brand,
                'discount_percent': discount_percent
            }
        except Exception as e:
            print(f"   ⚠ Помилка обробки JSON: {e}")
            return None
    
    def parse_page_html(self, html):
        """
        Парсинг сторінки через BeautifulSoup, використовуючи data-testid 
        для надійного вилучення цін, знижок та рейтингу.
        """
        soup = BeautifulSoup(html, 'html.parser')
        products_found = []
        
        # Шукаємо всі контейнери товарів
        product_items = soup.find_all('div', {'data-testid': 'productItem'})
        
        print(f"   (DEBUG: HTML: Знайдено {len(product_items)} блоків товарів для парсингу)")

        def clean_price(price_text_element):
            """Хелпер для очищення ціни від символів"""
            if price_text_element:
                text = price_text_element.get_text(strip=True)
                # Видаляємо всі не-цифрові символи
                return re.sub(r'[^\d]', '', text)
            return ""
        
        for container in product_items:
            try:
                # 1. Назва та URL
                name_elem = container.find('span', {'data-testid': 'productName'})
                title = name_elem.get('title', '') or name_elem.get_text(strip=True) if name_elem else ""
                
                link_elem = container.find('a', href=True)
                product_url = link_elem.get('href', '') if link_elem else ''
                url = f"{self.base_url}{product_url}" if product_url.startswith('/') else product_url

                # 2. Зображення
                image = ""
                img = container.find('img', {'data-testid': 'productImage'})
                if img:
                    image = img.get('src', '')
                
                # 3. Ціни та Знижки (використовуємо data-testid)
                price_elem = container.find('p', {'data-testid': 'finalPrice'})
                price = clean_price(price_elem)
                
                old_price_elem = container.find('p', {'data-testid': 'productFullPrice'})
                old_price = clean_price(old_price_elem)
                
                discount_percent_elem = container.find('span', {'data-testid': 'productDiscount'})
                discount_percent = re.sub(r'[^\d]', '', discount_percent_elem.get_text(strip=True)) if discount_percent_elem else ""
                
                # Сума знижки (розраховуємо)
                discount_amount_uah = ""
                if price and old_price:
                    try:
                        diff = int(old_price) - int(price)
                        discount_amount_uah = str(diff) if diff > 0 else ""
                    except ValueError:
                        pass

                # 4. Рейтинг та Відгуки
                stars = container.find_all('svg', {'data-testid': 'reviewStar'})
                rating = str(len(stars)) if stars else ""
                
                reviews = ""
                reviews_link = container.find('a', href=re.compile(r'#reviews'))
                if reviews_link:
                    reviews_p = reviews_link.find('p')
                    if reviews_p:
                        # Вилучаємо лише цифри
                        reviews = re.search(r'\d+', reviews_p.get_text(strip=True))
                        reviews = reviews.group(0) if reviews else ""

                # 5. Інші поля 
                availability = "В наявності"
                badges = ""
                brand = title.split()[0] if title else ""
                
                if discount_amount_uah or old_price:
                    print(f"   (DEBUG: HTML: Знижка: {discount_amount_uah} ₴ | Стара ціна: {old_price} ₴)")

                if title and price:
                    products_found.append({
                        'title': title,
                        'url': url,
                        'image': image,
                        'price': price,
                        'old_price': old_price,
                        'discount_amount_uah': discount_amount_uah,
                        'rating': rating,
                        'reviews_count': reviews,
                        'availability': availability,
                        'badges': badges,
                        'brand': brand,
                        'discount_percent': discount_percent
                    })
                    
            except Exception:
                continue
        
        return products_found

    ## 🔄 Методи керування парсингом
    
    def parse_file(self, filepath):
        """Парсинг одного локального файлу (гібридна логіка)"""
        html = self.read_local_page(filepath)
        if not html:
            return False
        
        page_products = []
        
        # 1. Спроба JSON
        json_products = self.extract_json_data(html)
        if json_products:
            print(f"✅ Знайдено {len(json_products)} товарів (JSON)")
            for product_data in json_products:
                product = self.parse_product_from_json(product_data)
                if product and product['title']:
                    page_products.append(product)
        
        # 2. Запасний варіант: HTML-парсинг, якщо JSON дав мало результатів
        if len(page_products) < 5:
            print(f"🔄 JSON-даних недостатньо або відсутні. Парсинг через HTML...")
            html_products = self.parse_page_html(html)
            
            # Якщо HTML дав більше результатів, ніж JSON, беремо HTML-результати
            if len(html_products) > len(page_products):
                 page_products = html_products
            
            if not page_products:
                print("⛔ Не вдалося знайти товари ні JSON, ні HTML.")
                return False

            print(f"✅ Знайдено {len(page_products)} товарів (ФІНАЛЬНО)")
        
        # Виведення результатів
        discount_count = 0
        for p in page_products:
            price_info = f"{p['price']} ₴"
            if p.get('discount_amount_uah') or p.get('old_price'):
                discount_count += 1
                if p.get('discount_amount_uah'):
                    price_info = f"Ціна: {p['price']} ₴ | Економія: {p['discount_amount_uah']} ₴"
                elif p.get('old_price'):
                    price_info = f"**{p['old_price']} ₴** → {p['price']} ₴"

            if p['discount_percent']:
                price_info += f" (-{p['discount_percent']}%)"
            
            print(f"   ✓ {p['title'][:55]}... | {price_info}")
        
        print(f"   (INFO: Знайдено {discount_count} акційних товарів у файлі.)")
        
        self.products.extend(page_products)
        return True
    
    def parse_all_local_files(self, folder_path):
        """Парсинг всіх HTML файлів у вказаній локальній папці"""
        
        # Отримуємо список HTML файлів і сортуємо їх
        html_files = sorted([f for f in os.listdir(folder_path) if f.endswith('.html')])
        total_files = len(html_files)
        
        print(f"\n{'='*80}")
        print(f"🚀 ПОЧАТОК ЛОКАЛЬНОГО ПАРСИНГУ")
        print(f"📂 Папка: {folder_path}")
        print(f"📄 Знайдено файлів HTML: {total_files}")
        print(f"{'='*80}\n")
        
        if total_files == 0:
            print("❌ У вказаній папці не знайдено жодного HTML-файлу.")
            return

        for index, filename in enumerate(html_files, 1):
            filepath = os.path.join(folder_path, filename)
            
            print(f"\n{'─'*80}")
            print(f"📄 Файл {index}/{total_files}: {filename}")
            print(f"{'─'*80}")
            
            self.parse_file(filepath)
            
            print(f"\n✅ Зібрано: {len(self.products)} товарів (загалом)")
        
        print(f"\n{'='*80}")
        print(f"🎉 ПАРСИНГ ФАЙЛІВ ЗАВЕРШЕНО!")
        print(f"{'='*80}")
        print(f"📊 Всього товарів: {len(self.products)}")

    ## 💾 Методи збереження
    
    def save_to_csv(self, filename='maudau_prybyrannia_ta_myttia.csv'):
        """Збереження у CSV"""
        if not self.products:
            print("❌ Немає даних")
            return
        
        # Використовуємо encoding='utf-8-sig' для коректного відображення кирилиці в Excel
        with open(filename, 'w', newline='', encoding='utf-8-sig') as f:
            fields = ['title', 'url', 'image', 'price', 'old_price', 'discount_amount_uah', 'rating',
                      'reviews_count', 'availability', 'badges', 'brand', 'discount_percent']
            writer = csv.DictWriter(f, fieldnames=fields)
            writer.writeheader()
            writer.writerows(self.products)
        
        print(f"\n✅ CSV: {filename}")
        print(f"📊 Товарів: {len(self.products)}")
    
    def save_to_excel(self, filename='maudau_prybyrannia_ta_myttia.xlsx'):
        """Збереження у Excel (потрібен 'openpyxl')"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Товари"
            
            headers = ['№', 'Назва', 'URL', 'Зображення', 'Ціна (актуальна)', 'Стара ціна', 'Сума знижки (₴)',
                       'Рейтинг', 'Відгуків', 'Наявність', 'Бейджі', 'Бренд', 'Знижка %']
            
            # Стилі для заголовків
            fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            font = Font(bold=True, color="FFFFFF", size=11)
            
            # Запис заголовків
            for col, h in enumerate(headers, 1):
                cell = ws.cell(1, col, h)
                cell.fill = fill
                cell.font = font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Запис даних
            for idx, p in enumerate(self.products, 2):
                ws.cell(idx, 1, idx-1)
                ws.cell(idx, 2, p['title'])
                ws.cell(idx, 3, p['url'])
                ws.cell(idx, 4, p['image'])
                ws.cell(idx, 5, p['price'])
                ws.cell(idx, 6, p.get('old_price', ''))
                ws.cell(idx, 7, p.get('discount_amount_uah', ''))
                ws.cell(idx, 8, p['rating'])
                ws.cell(idx, 9, p['reviews_count'])
                ws.cell(idx, 10, p['availability'])
                ws.cell(idx, 11, p['badges'])
                ws.cell(idx, 12, p['brand'])
                ws.cell(idx, 13, p['discount_percent'])
            
            # Автоматична ширина колонок
            for col in ws.columns:
                max_len = 0
                col_letter = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_len:
                            max_len = len(cell.value)
                    except:
                        pass
                # Обмежуємо ширину для зручності
                ws.column_dimensions[col_letter].width = min(max_len + 2, 60)
            
            wb.save(filename)
            print(f"✅ Excel: {filename}")
            
        except ImportError:
            print("⚠️   Встановіть openpyxl: pip install openpyxl")

    def print_sample(self, count=5):
        """Показати приклад перших товарів"""
        if not self.products:
            return
        
        print(f"\n{'='*80}")
        print(f"📋 ПРИКЛАД ТОВАРІВ ({min(count, len(self.products))} шт.):")
        print(f"{'='*80}\n")
        
        for i, p in enumerate(self.products[:count], 1):
            print(f"{i}. {p['title']}")
            print(f"   🏷️  Бренд: {p['brand']}")
            if p.get('old_price'):
                print(f"   💰 Стара: {p['old_price']} ₴")
            print(f"   💵 Ціна: {p['price']} ₴")
            if p.get('discount_amount_uah'):
                print(f"   💾 Економія: {p['discount_amount_uah']} ₴")
            if p['discount_percent']:
                print(f"   🔥 Знижка: -{p['discount_percent']}%")
            if p['rating']:
                print(f"   ⭐ Рейтинг: {p['rating']}/5")
            if p['reviews_count']:
                print(f"   💬 Відгуків: {p['reviews_count']}")
            if p['badges']:
                print(f"   🏆 Бейджі: {p['badges']}")
            print(f"   📦 {p['availability']}")
            print(f"   🔗 {p['url']}\n")


if __name__ == "__main__":
    
    # ❗️❗️❗️ ЗМІНІТЬ ЦЕЙ ШЛЯХ НА ШЛЯХ ДО ВАШОЇ ПАПКИ З HTML-ФАЙЛАМИ ❗️❗️❗️
    HTML_FOLDER_PATH = r'D:\testmaudau\maudau_prybyrannia_ta_myttia'
    # ❗️❗️❗️ ЗМІНІТЬ ЦЕЙ ШЛЯХ ДЛЯ ЗБЕРЕЖЕННЯ ❗️❗️❗️
    OUTPUT_FILE_BASE = r'D:\testmaudau'
    
    parser = MaudauParser()
    
    # Викликаємо метод для парсингу всіх локальних файлів у папці
    parser.parse_all_local_files(folder_path=HTML_FOLDER_PATH)
    
    # Приклад
    parser.print_sample(5)
    
    # Збереження
    parser.save_to_csv(f'{OUTPUT_FILE_BASE}.csv')
    parser.save_to_excel(f'{OUTPUT_FILE_BASE}.xlsx')
    
    print("\n✨ Готово!")